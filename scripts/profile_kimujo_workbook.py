#!/usr/bin/env python3
"""Print a compact profile of the Kimujo source workbook."""

from __future__ import annotations

import json
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


KEYWORDS = (
    "tenant",
    "name",
    "room",
    "house",
    "rent",
    "phone",
    "deposit",
    "payment",
    "paid",
    "balance",
    "outstanding",
    "arrears",
    "ebenezer",
    "kibuli",
    "nansana",
    "goshen",
)


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        text = " ".join(value.replace("\n", " ").split())
        return text[:120]
    return value


def row_values(ws, row_index):
    values = []
    for col_index in range(1, ws.max_column + 1):
        value = clean(ws.cell(row_index, col_index).value)
        if value is not None:
            values.append(f"{get_column_letter(col_index)}={value}")
    return values


def main() -> int:
    workbook_path = Path(sys.argv[1])
    wb_formula = load_workbook(workbook_path, data_only=False)
    wb_values = load_workbook(workbook_path, data_only=True)

    profile = {"workbook": str(workbook_path), "sheets": []}

    for sheet_name in wb_formula.sheetnames:
        ws_formula = wb_formula[sheet_name]
        ws_values = wb_values[sheet_name]
        nonempty_cells = 0
        formulas = []
        keyword_counts = Counter()
        numeric_cells = 0

        for row in ws_formula.iter_rows():
            for cell in row:
                value = cell.value
                if value is None:
                    continue
                nonempty_cells += 1
                if isinstance(value, str) and value.startswith("="):
                    if len(formulas) < 15:
                        formulas.append({"cell": cell.coordinate, "formula": value[:160]})
                    continue
                if isinstance(value, (int, float)):
                    numeric_cells += 1
                text = str(value).lower()
                for keyword in KEYWORDS:
                    if keyword in text:
                        keyword_counts[keyword] += 1

        nonempty_rows = []
        dense_rows = []
        for row_index in range(1, ws_values.max_row + 1):
            values = row_values(ws_values, row_index)
            if values:
                if len(nonempty_rows) < 30:
                    nonempty_rows.append({"row": row_index, "values": values[:24]})
                if len(values) >= 3 and len(dense_rows) < 20:
                    dense_rows.append({"row": row_index, "count": len(values), "values": values[:24]})

        profile["sheets"].append(
            {
                "name": sheet_name,
                "max_row": ws_formula.max_row,
                "max_column": ws_formula.max_column,
                "nonempty_cells": nonempty_cells,
                "numeric_cells": numeric_cells,
                "formula_count": sum(
                    1
                    for row in ws_formula.iter_rows()
                    for cell in row
                    if isinstance(cell.value, str) and cell.value.startswith("=")
                ),
                "formula_samples": formulas,
                "merged_range_count": len(ws_formula.merged_cells.ranges),
                "merged_range_samples": [str(rng) for rng in list(ws_formula.merged_cells.ranges)[:12]],
                "keyword_counts": dict(keyword_counts),
                "first_nonempty_rows": nonempty_rows,
                "first_dense_rows": dense_rows,
            }
        )

    print(json.dumps(profile, indent=2, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
