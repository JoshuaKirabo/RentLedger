#!/usr/bin/env python3
"""Dump compact row/column samples from a source workbook sheet."""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        return " ".join(value.replace("\n", " ").split())
    return value


def main() -> int:
    workbook_path = Path(sys.argv[1])
    sheet_name = sys.argv[2]
    start_row = int(sys.argv[3]) if len(sys.argv) > 3 else 1
    end_row = int(sys.argv[4]) if len(sys.argv) > 4 else start_row + 20

    wb = load_workbook(workbook_path, data_only=True)
    ws = wb[sheet_name]
    print(f"{ws.title}: rows={ws.max_row} cols={ws.max_column}")
    for row_index in range(start_row, min(end_row, ws.max_row) + 1):
        cells = []
        for col_index in range(1, ws.max_column + 1):
            value = clean(ws.cell(row_index, col_index).value)
            if value is not None:
                cells.append(f"{get_column_letter(col_index)}={value}")
        if cells:
            print(f"{row_index}: " + " | ".join(cells))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
