#!/usr/bin/env python3
"""Replace the seed SQLite data with tenant and payment rows from kimujo 2324."""

from __future__ import annotations

import json
import re
import sqlite3
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = Path("/Volumes/T7 Shield/kimujo 2324.xlsx")
DB_PATH = PROJECT_ROOT / "data" / "kimujjo_holdings_database.db"
SQL_PATH = PROJECT_ROOT / "sql" / "kimujo_property_management.sql"
REPORT_PATH = PROJECT_ROOT / "data" / "import_kimujo_2324_report.json"
BACKUP_DIR = PROJECT_ROOT / "data" / "backups"

MIN_RENT = 100_000
MAX_RENT = 15_000_000
REGULAR_RENT_MAX = 5_000_000

MONTHS = (
    "jan",
    "feb",
    "mar",
    "apr",
    "may",
    "jun",
    "jul",
    "aug",
    "sep",
    "oct",
    "nov",
    "dec",
)

SKIP_NAMES = {"total", "totals", "0", "do"}


@dataclass(frozen=True)
class MonthColumn:
    rent_month: str
    amount_col: int
    receipt_col: int


@dataclass(frozen=True)
class SheetConfig:
    sheet: str
    estate: str
    code: str
    name_col: str
    phone_col: str | None
    unit_col: str | None
    security_col: str | None
    security_receipt_col: str | None
    first_row: int
    last_row: int
    sections: tuple[tuple[int, str, int], ...]


@dataclass
class PaymentEvent:
    rent_month: str
    amount: int
    receipt_raw: str | None
    amount_cell: str
    receipt_cell: str


@dataclass
class TenantSource:
    source_id: str
    sheet: str
    estate: str
    estate_code: str
    row: int
    raw_name: str
    first_name: str
    middle_name: str | None
    last_name: str
    raw_phone: str | None
    phone_number: str
    generated_phone: bool
    unit_number: str
    security_amount: int | None
    security_receipt: str | None
    monthly_rent: int
    start_date: str
    latest_payment_date: str | None
    is_active: bool = True
    end_date: str | None = None
    payment_events: list[PaymentEvent] = field(default_factory=list)


def col(letter: str) -> int:
    return column_index_from_string(letter)


def month_columns(year: int, start_col: str, count: int) -> tuple[MonthColumn, ...]:
    start = col(start_col)
    return tuple(
        MonthColumn(f"{year}-{month_index + 1:02d}", start + (month_index * 2), start + (month_index * 2) + 1)
        for month_index in range(count)
    )


SHEETS: tuple[SheetConfig, ...] = (
    SheetConfig(
        sheet="Moriah Kikaya",
        estate="Moriah Kikaya",
        code="MOR",
        name_col="A",
        phone_col="C",
        unit_col="B",
        security_col="D",
        security_receipt_col="E",
        first_row=7,
        last_row=54,
        sections=((2023, "F", 12), (2024, "AF", 12), (2025, "BE", 12), (2026, "CC", 6)),
    ),
    SheetConfig(
        sheet="Horeb Ntinda",
        estate="Horeb Ntinda",
        code="HOR",
        name_col="A",
        phone_col="B",
        unit_col="C",
        security_col="D",
        security_receipt_col=None,
        first_row=4,
        last_row=16,
        sections=((2023, "E", 12), (2024, "AD", 12), (2025, "BB", 12), (2026, "BZ", 3)),
    ),
    SheetConfig(
        sheet="Kibuli",
        estate="Kibuli",
        code="KIB",
        name_col="A",
        phone_col="B",
        unit_col="C",
        security_col="D",
        security_receipt_col=None,
        first_row=4,
        last_row=20,
        sections=((2023, "E", 12), (2024, "AD", 12), (2025, "BC", 5)),
    ),
    SheetConfig(
        sheet="Nansana",
        estate="Nansana",
        code="NAN",
        name_col="A",
        phone_col="B",
        unit_col="C",
        security_col="D",
        security_receipt_col=None,
        first_row=4,
        last_row=22,
        sections=((2023, "E", 12), (2024, "AD", 12), (2025, "BB", 9)),
    ),
    SheetConfig(
        sheet="Salem Maganjo",
        estate="Salem Maganjo",
        code="SAL",
        name_col="A",
        phone_col="B",
        unit_col="C",
        security_col="D",
        security_receipt_col=None,
        first_row=5,
        last_row=21,
        sections=((2023, "E", 12), (2024, "AD", 12), (2025, "BB", 5)),
    ),
    SheetConfig(
        sheet="Ebenezer Kawempe",
        estate="Ebenezer Kawempe",
        code="EBE",
        name_col="A",
        phone_col="B",
        unit_col="C",
        security_col="D",
        security_receipt_col=None,
        first_row=3,
        last_row=27,
        sections=((2023, "E", 12), (2024, "AD", 12), (2025, "BB", 12), (2026, "CA", 5)),
    ),
    SheetConfig(
        sheet="Goshen Kawempe",
        estate="Goshen Kawempe",
        code="GOS",
        name_col="A",
        phone_col="B",
        unit_col="D",
        security_col="E",
        security_receipt_col=None,
        first_row=3,
        last_row=25,
        sections=((2023, "F", 12), (2024, "AE", 12), (2025, "BD", 7)),
    ),
)


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = " ".join(str(value).replace("\n", " ").split()).strip()
    return text or None


def source_value(ws_values, ws_formulas, row: int, col_index: int) -> Any:
    value = ws_values.cell(row, col_index).value
    if value is not None:
        return value
    formula = ws_formulas.cell(row, col_index).value
    return formula


def parse_amount(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        amount = int(round(float(value)))
        return amount if amount > 0 else None

    text = clean_text(value)
    if not text:
        return None
    text = text.replace(",", "").replace(" ", "")
    if text.startswith("="):
        text = text[1:]
    if re.fullmatch(r"\d+(?:\.\d+)?", text):
        amount = int(round(float(text)))
        return amount if amount > 0 else None
    if re.fullmatch(r"\d+(?:\.\d+)?(?:\+\d+(?:\.\d+)?)+", text):
        amount = int(round(sum(float(part) for part in text.split("+"))))
        return amount if amount > 0 else None
    return None


def clean_receipt(value: Any) -> str | None:
    text = clean_text(value)
    if not text:
        return None
    lowered = text.lower()
    if lowered in {"0", "no", "none", "nil", "nan"} or "no rct" in lowered:
        return None
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    return text


def looks_like_phone(value: str | None) -> bool:
    if not value:
        return False
    digits = re.sub(r"\D", "", value)
    return len(digits) >= 7 and (digits.startswith("0") or digits.startswith("7") or digits.startswith("256"))


def normalize_unit(value: Any) -> str | None:
    text = clean_text(value)
    if not text:
        return None
    return re.sub(r"\s+", " ", text).strip().upper()


NAME_OVERRIDES: dict[str, tuple[str, str | None, str]] = {
    "muhammedwaswa": ("Muhammed", None, "Waswa"),
    "michealkakule": ("Michael", None, "Kakule"),
}

BUSINESS_TENANT_NAMES: dict[str, str] = {
    "twebacoltd": "Tweba Co Ltd",
    "opportunitybank": "Opportunity Bank",
    "ugachiccck": "Ugachick",
    "ugachick": "Ugachick",
    "godsloverecords": "God's Love Records",
    "medorapharma": "Medora Pharma",
}


def normalize_name_part(value: str) -> str:
    text = re.sub(r"\s+", " ", value.strip())
    if not text:
        return "Unknown"
    return text[0].upper() + text[1:].lower()


def preprocess_name(raw_name: str) -> str:
    text = re.sub(r"\s+", " ", raw_name.strip())
    text = re.sub(r"([a-z])([A-Z])", r"\1 \2", text)
    text = re.sub(r"([A-Z]+)([A-Z][a-z])", r"\1 \2", text)
    return text


def split_name(raw_name: str) -> tuple[str, str | None, str]:
    override_key = re.sub(r"\s+", "", raw_name.strip().lower())
    if override_key in NAME_OVERRIDES:
        return NAME_OVERRIDES[override_key]

    cleaned = preprocess_name(raw_name)
    parts = [part for part in cleaned.split(" ") if part]
    if not parts:
        return "Unknown", None, "Unknown"
    if len(parts) == 1:
        return normalize_name_part(parts[0]), None, "Unknown"
    if len(parts) == 2:
        return normalize_name_part(parts[0]), None, normalize_name_part(parts[1])
    return (
        normalize_name_part(parts[0]),
        normalize_name_part(" ".join(parts[1:-1])),
        normalize_name_part(parts[-1]),
    )


def valid_name(raw_name: str | None) -> bool:
    if not raw_name:
        return False
    lowered = raw_name.strip().lower()
    if lowered in SKIP_NAMES:
        return False
    return bool(re.search(r"[A-Za-z0-9]", raw_name))


class PhoneNormalizer:
    def __init__(self) -> None:
        self.used: set[str] = set()
        self.generated_counter = 1
        self.generated: list[dict[str, Any]] = []

    def _candidate_numbers(self, raw_phone: str | None) -> list[str]:
        if not raw_phone:
            return []
        groups = re.findall(r"\d{7,}", raw_phone)
        candidates: list[str] = []
        for group in groups:
            if group.startswith("2567") and len(group) >= 12:
                candidates.append(f"+{group[:12]}")
            if group.startswith("07") and len(group) >= 10:
                candidates.append(f"+256{group[1:10]}")
            if group.startswith("7") and len(group) >= 9:
                candidates.append(f"+256{group[:9]}")
        return candidates

    def normalize(self, raw_phone: str | None, source_id: str, reason: str = "") -> tuple[str, bool]:
        for candidate in self._candidate_numbers(raw_phone):
            if re.fullmatch(r"\+2567\d{8}", candidate) and candidate not in self.used:
                self.used.add(candidate)
                return candidate, False

        while True:
            local = f"799{self.generated_counter:06d}"
            self.generated_counter += 1
            candidate = f"+256{local}"
            if candidate not in self.used:
                self.used.add(candidate)
                self.generated.append(
                    {
                        "source_id": source_id,
                        "raw_phone": raw_phone,
                        "generated_phone": candidate,
                        "reason": reason or "missing, invalid, or duplicate phone",
                    }
                )
                return candidate, True


def infer_monthly_rent(payment_events: list[PaymentEvent], security_amount: int | None) -> int:
    candidates = [event.amount for event in payment_events if MIN_RENT <= event.amount <= REGULAR_RENT_MAX]
    if candidates:
        counts = Counter(candidates)
        return counts.most_common(1)[0][0]
    if security_amount and MIN_RENT <= security_amount <= MAX_RENT:
        return security_amount
    return 300_000


def clamp_rent(amount: int) -> int:
    return max(MIN_RENT, min(MAX_RENT, int(amount)))


def date_for_month(rent_month: str) -> str:
    return f"{rent_month}-01"


def due_date_for_month(rent_month: str) -> str:
    return f"{rent_month}-01"


def build_reference(prefix: str, raw: str | None, fallback: str, used: set[str]) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9+/_-]+", "-", raw or "").strip("-")
    base = f"{prefix}-{cleaned}" if cleaned else fallback
    base = base[:120]
    candidate = base
    suffix = 2
    while candidate in used:
        extra = f"-{suffix}"
        candidate = f"{base[:120 - len(extra)]}{extra}"
        suffix += 1
    used.add(candidate)
    return candidate


def receipt_number(counter: int) -> str:
    return str(counter).zfill(6)


def load_schema_ddl() -> str:
    return SQL_PATH.read_text()


def backup_database() -> str | None:
    if not DB_PATH.exists():
        return None
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    backup_path = BACKUP_DIR / f"kimujjo_holdings_database.before-real-import-{datetime.now():%Y%m%d-%H%M%S}.db"
    source = sqlite3.connect(DB_PATH)
    try:
        dest = sqlite3.connect(backup_path)
        try:
            source.backup(dest)
        finally:
            dest.close()
    finally:
        source.close()
    return str(backup_path)


def collect_sources(workbook_path: Path) -> tuple[list[TenantSource], dict[str, Any]]:
    wb_values = load_workbook(workbook_path, data_only=True)
    wb_formulas = load_workbook(workbook_path, data_only=False)
    phone_normalizer = PhoneNormalizer()
    tenants: list[TenantSource] = []
    skipped_rows: list[dict[str, Any]] = []

    for config in SHEETS:
        ws_values = wb_values[config.sheet]
        ws_formulas = wb_formulas[config.sheet]
        month_map = [item for year, start, count in config.sections for item in month_columns(year, start, count)]

        for row in range(config.first_row, config.last_row + 1):
            source_id = f"{config.sheet}!{row}"
            raw_name = clean_text(source_value(ws_values, ws_formulas, row, col(config.name_col)))
            if not valid_name(raw_name):
                continue

            raw_phone = clean_text(source_value(ws_values, ws_formulas, row, col(config.phone_col))) if config.phone_col else None
            unit_raw = clean_text(source_value(ws_values, ws_formulas, row, col(config.unit_col))) if config.unit_col else None
            if not raw_phone and looks_like_phone(unit_raw):
                raw_phone = unit_raw
                unit_raw = None
            unit_number = normalize_unit(unit_raw)

            payment_events: list[PaymentEvent] = []
            for month_col in month_map:
                amount_value = source_value(ws_values, ws_formulas, row, month_col.amount_col)
                amount = parse_amount(amount_value)
                if not amount:
                    continue
                receipt_value = source_value(ws_values, ws_formulas, row, month_col.receipt_col)
                payment_events.append(
                    PaymentEvent(
                        rent_month=month_col.rent_month,
                        amount=amount,
                        receipt_raw=clean_receipt(receipt_value),
                        amount_cell=f"{get_column_letter(month_col.amount_col)}{row}",
                        receipt_cell=f"{get_column_letter(month_col.receipt_col)}{row}",
                    )
                )

            security_amount = (
                parse_amount(source_value(ws_values, ws_formulas, row, col(config.security_col)))
                if config.security_col
                else None
            )
            security_receipt = (
                clean_receipt(source_value(ws_values, ws_formulas, row, col(config.security_receipt_col)))
                if config.security_receipt_col
                else None
            )

            if not unit_number and not payment_events and not security_amount:
                skipped_rows.append(
                    {"source_id": source_id, "name": raw_name, "reason": "no unit, security amount, or payment data"}
                )
                continue
            if not unit_number:
                unit_number = f"UNASSIGNED {row}"

            phone_number, generated_phone = phone_normalizer.normalize(raw_phone, source_id)
            first_name, middle_name, last_name = split_name(raw_name or "Unknown")
            monthly_rent = clamp_rent(infer_monthly_rent(payment_events, security_amount))
            start_month = min((event.rent_month for event in payment_events), default="2023-01")
            latest_month = max((event.rent_month for event in payment_events), default=None)

            tenants.append(
                TenantSource(
                    source_id=source_id,
                    sheet=config.sheet,
                    estate=config.estate,
                    estate_code=config.code,
                    row=row,
                    raw_name=raw_name or "Unknown",
                    first_name=first_name,
                    middle_name=middle_name,
                    last_name=last_name,
                    raw_phone=raw_phone,
                    phone_number=phone_number,
                    generated_phone=generated_phone,
                    unit_number=unit_number,
                    security_amount=security_amount,
                    security_receipt=security_receipt,
                    monthly_rent=monthly_rent,
                    start_date=date_for_month(start_month),
                    latest_payment_date=date_for_month(latest_month) if latest_month else None,
                    payment_events=payment_events,
                )
            )

    mark_active_tenants(tenants)
    report = {
        "skipped_rows": skipped_rows,
        "generated_phones": phone_normalizer.generated,
    }
    return tenants, report


def mark_active_tenants(tenants: list[TenantSource]) -> None:
    grouped: dict[tuple[str, str], list[TenantSource]] = defaultdict(list)
    for tenant in tenants:
        grouped[(tenant.estate, tenant.unit_number)].append(tenant)

    for group in grouped.values():
        if len(group) == 1:
            group[0].is_active = True
            group[0].end_date = None
            continue

        def sort_key(tenant: TenantSource) -> tuple[str, int, int]:
            latest = tenant.latest_payment_date or tenant.start_date
            return latest, len(tenant.payment_events), tenant.row

        active = max(group, key=sort_key)
        for tenant in group:
            tenant.is_active = tenant is active
            tenant.end_date = None if tenant is active else (tenant.latest_payment_date or tenant.start_date)


def insert_real_data(tenants: list[TenantSource]) -> dict[str, Any]:
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON")
    conn.executescript(load_schema_ddl())

    estate_ids: dict[str, int] = {}
    for config in SHEETS:
        cursor = conn.execute(
            "INSERT INTO estates (estate_code, estate_name) VALUES (?, ?)",
            (config.code, config.estate),
        )
        estate_ids[config.estate] = int(cursor.lastrowid)

    unit_ids: dict[tuple[str, str], int] = {}
    unit_rents: dict[tuple[str, str], int] = {}
    for tenant in tenants:
        key = (tenant.estate, tenant.unit_number)
        current = unit_rents.get(key, 0)
        if tenant.is_active or not current:
            unit_rents[key] = tenant.monthly_rent

    for (estate, unit_number), rent in sorted(unit_rents.items()):
        cursor = conn.execute(
            "INSERT INTO units (estate_id, unit_number, listed_monthly_rent) VALUES (?, ?, ?)",
            (estate_ids[estate], unit_number, clamp_rent(rent)),
        )
        unit_ids[(estate, unit_number)] = int(cursor.lastrowid)

    tenant_ids: dict[str, int] = {}
    tenancy_ids: dict[str, int] = {}
    security_deposit_ids: dict[str, int] = {}

    for tenant in tenants:
        cursor = conn.execute(
            """
            INSERT INTO tenants (first_name, middle_name, last_name, phone_number, is_active)
            VALUES (?, ?, ?, ?, ?)
            """,
            (
                tenant.first_name,
                tenant.middle_name,
                tenant.last_name,
                tenant.phone_number,
                1 if tenant.is_active else 0,
            ),
        )
        tenant_id = int(cursor.lastrowid)
        tenant_ids[tenant.source_id] = tenant_id

        cursor = conn.execute(
            """
            INSERT INTO tenancy_assignments (
                tenant_id,
                unit_id,
                start_date,
                end_date,
                agreed_monthly_rent
            ) VALUES (?, ?, ?, ?, ?)
            """,
            (
                tenant_id,
                unit_ids[(tenant.estate, tenant.unit_number)],
                tenant.start_date,
                tenant.end_date,
                tenant.monthly_rent,
            ),
        )
        tenancy_id = int(cursor.lastrowid)
        tenancy_ids[tenant.source_id] = tenancy_id

        expected_deposit = tenant.security_amount or tenant.monthly_rent
        cursor = conn.execute(
            "INSERT INTO security_deposits (tenancy_id, expected_amount) VALUES (?, ?)",
            (tenancy_id, max(1, expected_deposit)),
        )
        security_deposit_ids[tenant.source_id] = int(cursor.lastrowid)

    used_payment_refs: set[str] = set()
    receipt_counter = 1
    security_payment_count = 0
    for tenant in tenants:
        if not tenant.security_amount:
            continue
        tenant_id = tenant_ids[tenant.source_id]
        security_deposit_id = security_deposit_ids[tenant.source_id]
        fallback = f"SEC-{tenant.estate_code}-{tenant.row}"
        payment_ref = build_reference("SEC", tenant.security_receipt, fallback, used_payment_refs)
        cursor = conn.execute(
            """
            INSERT INTO payments (
                tenant_id,
                payment_reference,
                payment_type,
                amount,
                payment_date,
                payment_method,
                payment_status
            ) VALUES (?, ?, 'SECURITY_DEPOSIT', ?, ?, 'CASH', 'POSTED')
            """,
            (tenant_id, payment_ref, tenant.security_amount, tenant.start_date),
        )
        payment_id = int(cursor.lastrowid)
        conn.execute(
            """
            INSERT INTO security_deposit_payments (
                security_deposit_id,
                payment_id,
                allocated_amount
            ) VALUES (?, ?, ?)
            """,
            (security_deposit_id, payment_id, tenant.security_amount),
        )
        conn.execute(
            "INSERT INTO receipts (payment_id, receipt_number, issued_at, issued_by) VALUES (?, ?, ?, 'IMPORT')",
            (payment_id, receipt_number(receipt_counter), f"{tenant.start_date}T09:00:00"),
        )
        receipt_counter += 1
        security_payment_count += 1

    obligation_ids: dict[tuple[str, str], int] = {}
    obligation_totals: dict[tuple[str, str], int] = defaultdict(int)
    for tenant in tenants:
        for event in tenant.payment_events:
            obligation_totals[(tenant.source_id, event.rent_month)] += event.amount

    for (source_id, rent_month), amount_due in sorted(obligation_totals.items()):
        cursor = conn.execute(
            """
            INSERT INTO rent_obligations (
                tenancy_id,
                rent_month,
                due_date,
                amount_due
            ) VALUES (?, ?, ?, ?)
            """,
            (tenancy_ids[source_id], rent_month, due_date_for_month(rent_month), amount_due),
        )
        obligation_ids[(source_id, rent_month)] = int(cursor.lastrowid)

    rent_payment_count = 0
    allocation_count = 0
    for tenant in tenants:
        groups: dict[tuple[str, str], list[PaymentEvent]] = defaultdict(list)
        for event in tenant.payment_events:
            if event.receipt_raw:
                group_key = ("receipt", event.receipt_raw)
            else:
                group_key = ("cell", event.amount_cell)
            groups[group_key].append(event)

        for (group_kind, raw_ref), events in sorted(groups.items(), key=lambda item: min(e.rent_month for e in item[1])):
            amount = sum(event.amount for event in events)
            first_month = min(event.rent_month for event in events)
            fallback = f"PAY-{tenant.estate_code}-{tenant.row}-{first_month}-{events[0].amount_cell}"
            payment_ref = build_reference("RCP", raw_ref if group_kind == "receipt" else None, fallback, used_payment_refs)
            cursor = conn.execute(
                """
                INSERT INTO payments (
                    tenant_id,
                    payment_reference,
                    payment_type,
                    amount,
                    payment_date,
                    payment_method,
                    payment_status
                ) VALUES (?, ?, 'RENT', ?, ?, 'CASH', 'POSTED')
                """,
                (tenant_ids[tenant.source_id], payment_ref, amount, date_for_month(first_month)),
            )
            payment_id = int(cursor.lastrowid)

            month_allocations: dict[str, int] = defaultdict(int)
            for event in events:
                month_allocations[event.rent_month] += event.amount
            for rent_month, allocated_amount in sorted(month_allocations.items()):
                conn.execute(
                    """
                    INSERT INTO payment_allocations (
                        payment_id,
                        rent_obligation_id,
                        allocated_amount
                    ) VALUES (?, ?, ?)
                    """,
                    (payment_id, obligation_ids[(tenant.source_id, rent_month)], allocated_amount),
                )
                allocation_count += 1

            conn.execute(
                "INSERT INTO receipts (payment_id, receipt_number, issued_at, issued_by) VALUES (?, ?, ?, 'IMPORT')",
                (payment_id, receipt_number(receipt_counter), f"{date_for_month(first_month)}T09:00:00"),
            )
            receipt_counter += 1
            rent_payment_count += 1

    conn.commit()

    counts = {
        name: conn.execute(f"SELECT COUNT(*) FROM {name}").fetchone()[0]
        for name in (
            "estates",
            "units",
            "tenants",
            "tenancy_assignments",
            "security_deposits",
            "rent_obligations",
            "payments",
            "payment_allocations",
            "security_deposit_payments",
            "receipts",
        )
    }
    active_tenants = conn.execute("SELECT COUNT(*) FROM tenants WHERE is_active = 1").fetchone()[0]
    conn.close()

    return {
        "counts": counts,
        "active_tenants": active_tenants,
        "inactive_tenants": len(tenants) - active_tenants,
        "rent_payment_count": rent_payment_count,
        "security_payment_count": security_payment_count,
        "allocation_count": allocation_count,
    }


def main() -> int:
    workbook_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_WORKBOOK
    if not workbook_path.exists():
        raise FileNotFoundError(workbook_path)

    tenants, report = collect_sources(workbook_path)
    backup_path = backup_database()
    import_summary = insert_real_data(tenants)

    duplicate_units = [
        {
            "estate": estate,
            "unit": unit,
            "tenants": [
                {
                    "source_id": tenant.source_id,
                    "name": tenant.raw_name,
                    "active": tenant.is_active,
                    "latest_payment_date": tenant.latest_payment_date,
                }
                for tenant in group
            ],
        }
        for (estate, unit), group in sorted(group_sources_by_unit(tenants).items())
        if len(group) > 1
    ]

    final_report = {
        "source_workbook": str(workbook_path),
        "database": str(DB_PATH),
        "backup_database": backup_path,
        "imported_at": datetime.now().isoformat(timespec="seconds"),
        **import_summary,
        "duplicate_unit_resolution": duplicate_units,
        **report,
    }
    REPORT_PATH.write_text(json.dumps(final_report, indent=2))
    print(json.dumps(final_report, indent=2))
    return 0


def group_sources_by_unit(tenants: list[TenantSource]) -> dict[tuple[str, str], list[TenantSource]]:
    grouped: dict[tuple[str, str], list[TenantSource]] = defaultdict(list)
    for tenant in tenants:
        grouped[(tenant.estate, tenant.unit_number)].append(tenant)
    return grouped


if __name__ == "__main__":
    raise SystemExit(main())
